"""
Office Open XML Agile Encryption Decryptor (ECMA-376 spec)
Dùng cryptography (có sẵn trên mọi Python) + stdlib
"""
import io, struct, hashlib, base64, xml.etree.ElementTree as ET
import olefile

try:
    from cryptography.hazmat.primitives.ciphers import Cipher, algorithms, modes
    from cryptography.hazmat.backends import default_backend
    HAS_CRYPTO = True
except ImportError:
    HAS_CRYPTO = False

BLK_VI = bytes([0xFE, 0xA7, 0xD2, 0x76, 0x3B, 0x4B, 0x9E, 0x79])
BLK_VH = bytes([0xD7, 0xAA, 0x0F, 0x6D, 0x30, 0x61, 0x34, 0x4E])
BLK_EK = bytes([0x14, 0x6E, 0x0B, 0xE7, 0xAB, 0xAC, 0xD0, 0xD6])

NS = 'http://schemas.microsoft.com/office/2006/encryption'
PNS = 'http://schemas.microsoft.com/office/2006/keyEncryptor/password'


def aes_cbc_decrypt(key, iv, data):
    """AES-CBC decrypt using cryptography library"""
    cipher = Cipher(algorithms.AES(key), modes.CBC(iv), backend=default_backend())
    decryptor = cipher.decryptor()
    return decryptor.update(data) + decryptor.finalize()


def decrypt_agile(file_bytes, password):
    if not HAS_CRYPTO:
        raise ImportError("cryptography required: pip install cryptography")

    ole = olefile.OleFileIO(io.BytesIO(file_bytes))
    enc_info = ole.openstream('EncryptionInfo').read()
    root = ET.fromstring(enc_info[8:])

    kd = root.find(f'./{{{NS}}}keyData')
    data_salt = base64.b64decode(kd.get('saltValue'))

    ek = root.find(f'.//{{{NS}}}keyEncryptors/{{{NS}}}keyEncryptor[@uri="{PNS}"]/{{{PNS}}}encryptedKey')
    spin = int(ek.get('spinCount'))
    ek_salt = base64.b64decode(ek.get('saltValue'))
    verifier_input  = base64.b64decode(ek.get('encryptedVerifierHashInput'))
    verifier_hash   = base64.b64decode(ek.get('encryptedVerifierHashValue'))
    encrypted_key   = base64.b64decode(ek.get('encryptedKeyValue'))

    # Iterated hash
    h = hashlib.sha512(ek_salt + password.encode('utf-16-le'))
    for i in range(0, spin):
        h = hashlib.sha512(struct.pack('<I', i) + h.digest())
    h_digest = h.digest()

    # Verify password
    k1 = hashlib.sha512(h_digest + BLK_VI).digest()[:32]
    verifier = aes_cbc_decrypt(k1, ek_salt, verifier_input)[:16]

    k2 = hashlib.sha512(h_digest + BLK_VH).digest()[:32]
    decrypted_hash = aes_cbc_decrypt(k2, ek_salt, verifier_hash)[:64]

    if hashlib.sha512(verifier).digest() != decrypted_hash:
        raise ValueError("Wrong password!")

    # Decrypt secret key
    k3 = hashlib.sha512(h_digest + BLK_EK).digest()[:32]
    secret_key = aes_cbc_decrypt(k3, ek_salt, encrypted_key)[:32]

    # Decrypt EncryptedPackage (4096-byte segments)
    enc_pkg = ole.openstream('EncryptedPackage').read()
    total_size = struct.unpack_from('<Q', enc_pkg, 0)[0]

    result = io.BytesIO()
    remaining = total_size
    offset = 8

    for seg_idx in range(0, (total_size + 4095) // 4096):
        seg_size = min(4096, remaining)
        padded_size = ((seg_size + 15) // 16) * 16
        seg_data = enc_pkg[offset:offset + padded_size]
        seg_iv = hashlib.sha512(data_salt + struct.pack('<I', seg_idx)).digest()[:16]
        decrypted = aes_cbc_decrypt(secret_key, seg_iv, seg_data)
        result.write(decrypted[:seg_size])
        remaining -= seg_size
        offset += padded_size
        if remaining <= 0:
            break

    ole.close()
    return result.getvalue()


def decrypt_excel(file_bytes, password='sp'):
    import openpyxl
    try:
        openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True).close()
        return file_bytes
    except:
        pass
    return decrypt_agile(file_bytes, password)


if __name__ == '__main__':
    import sys, openpyxl
    src = sys.argv[1] if len(sys.argv) > 1 else 'test.xlsx'
    dst = sys.argv[2] if len(sys.argv) > 2 else src.replace('.xlsx', '_DECRYPTED.xlsx')
    pw = sys.argv[3] if len(sys.argv) > 3 else 'sp'

    with open(src, 'rb') as f:
        data = f.read()
    decrypted = decrypt_excel(data, pw)
    with open(dst, 'wb') as f:
        f.write(decrypted)

    wb = openpyxl.load_workbook(io.BytesIO(decrypted), data_only=True)
    print(f"✅ Decrypted OK! {len(wb.sheetnames)} sheets → {dst}")
    wb.close()
