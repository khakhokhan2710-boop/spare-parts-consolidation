# HOYA Spare Parts Consolidation - Flask Web App
import os
import io
import tempfile
import shutil
from pathlib import Path
from flask import Flask, request, jsonify, send_file, render_template_string

import msoffcrypto
import openpyxl

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB
app.config['UPLOAD_FOLDER'] = tempfile.mkdtemp()

# ============================================================
# MAP
# ============================================================
SAMPLE_SHEET_MAP = {
    'AO': 'AO', 'AR': 'AR COAT', 'AS': 'Assembling',
    'FI': 'Final', 'GK': 'GK', 'HC': 'HARD COAT',
    'HR': 'HOLV ADMIN COMMON', 'IT': 'IT',
    'Material WH': 'Material WH', 'MF': 'Mixing Filling',
    'MF-174': 'Mixing Filling 174', 'MF-PNX': 'Mixing Filling PNX',
    'Mold': 'Mold Preparation', 'PRO PLAN': 'PRO PLAN',
    'Pro-WH': 'PRO WH', 'QA': 'QA', 'Sensity': 'SUNTECH',
    'Hóa chất-Technical': 'TECHNICAL', 'Facility': 'TPM',
    'Visual': 'Visual', 'Export': 'EXPORT',
}

MONTHS = ['Mar', 'Apr', 'May', 'Jun']
SAMPLE_OUTPUT_COL = 13   # 0-indexed: N
MASTER_153_OUTPUT_COL = 22  # 0-indexed: W
MASTER_152_OUTPUT_COL = 42  # 0-indexed: AP


# ============================================================
# DECRYPT
# ============================================================
def decrypt_excel(file_bytes, password='sp'):
    """Giải mã file Excel có password"""
    try:
        # Thử đọc trực tiếp (không password)
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        wb.close()
        return file_bytes  # Không cần giải mã
    except Exception:
        pass

    # Cần giải mã
    try:
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_in:
            tmp_in.write(file_bytes)
            tmp_in_path = tmp_in.name

        with open(tmp_in_path, 'rb') as f:
            file = msoffcrypto.OfficeFile(f)
            file.load_key(password=password)
            out_bytes = io.BytesIO()
            file.decrypt(out_bytes)
            out_bytes.seek(0)

        os.unlink(tmp_in_path)
        return out_bytes.getvalue()
    except Exception as e:
        raise ValueError(f"Không thể giải mã file (sai password?): {e}")


# ============================================================
# READ SAMPLE
# ============================================================
def read_sample_data(file_bytes):
    """Đọc cột N từ file sample"""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    results = {}

    for sn in wb.sheetnames:
        if sn == 'Check':
            continue
        ws = wb[sn]
        section = SAMPLE_SHEET_MAP.get(sn, sn)
        rows = []

        for r in range(7, ws.max_row + 1):
            qty = ws.cell(row=r, column=SAMPLE_OUTPUT_COL + 1).value  # 1-indexed
            if qty is None or str(qty).strip() in ('', '0', 0):
                continue
            try:
                qty = float(qty)
            except (ValueError, TypeError):
                continue

            desc = str(ws.cell(row=r, column=5).value or '').strip()
            po = str(ws.cell(row=r, column=6).value or '').strip()
            invoice = str(ws.cell(row=r, column=3).value or '').strip()

            if desc.upper() == 'TOTAL' or po.upper() == 'TOTAL':
                continue

            rows.append({'desc': desc, 'po': po, 'invoice': invoice, 'qty': qty})

        if rows:
            results[section] = rows

    wb.close()
    return results


# ============================================================
# MATCHING
# ============================================================
def normalize(s):
    import re
    return re.sub(r'[^a-z0-9]', '', str(s).lower().strip())

def match_score(sample, master):
    score = 0
    poS = normalize(sample['po'])
    poM = normalize(master['po'])
    if poS and poM and poS == poM:
        score += 10
    elif poS and poM and (poS in poM or poM in poS):
        score += 5

    descS = normalize(sample['desc'])
    descM = normalize(master['desc'])
    if descS and descM and descS == descM:
        score += 8
    elif descS and descM and (descS in descM or descM in descS):
        score += 4

    return score


def consolidate(master_bytes, sample_data, month):
    """Match sample vào master, trả về (output_bytes, stats)"""
    wb = openpyxl.load_workbook(io.BytesIO(master_bytes), data_only=False)
    wb_data = openpyxl.load_workbook(io.BytesIO(master_bytes), data_only=True)

    s153 = f"153-{month}'26"
    s152 = f"152-{month}'26"

    matched153 = 0
    matched152 = 0
    unmatched = []

    for section, rows in sample_data.items():
        # Match 153
        if s153 in wb.sheetnames:
            ws = wb[s153]
            ws_d = wb_data[s153]

            masters = []
            for r in range(12, ws_d.max_row + 1):
                ms = str(ws_d.cell(row=r, column=2).value or '').strip().upper()
                if ms != section.upper():
                    continue
                masters.append({
                    'row': r,
                    'desc': str(ws_d.cell(row=r, column=1).value or ''),
                    'po': str(ws_d.cell(row=r, column=7).value or ''),
                })

            used = set()
            for row in rows:
                best = {'score': 0, 'mr': None}
                for mr in masters:
                    if mr['row'] in used:
                        continue
                    s = match_score(row, mr)
                    if s > best['score']:
                        best = {'score': s, 'mr': mr}
                if best['score'] >= 4 and best['mr']:
                    ws.cell(row=best['mr']['row'], column=MASTER_153_OUTPUT_COL + 1).value = row['qty']
                    used.add(best['mr']['row'])
                    matched153 += 1
                else:
                    unmatched.append({'section': section, 'po': row['po'], 'desc': row['desc'][:60], 'qty': row['qty']})

        # Match 152
        if s152 in wb.sheetnames:
            ws = wb[s152]
            ws_d = wb_data[s152]
            masters = []
            for r in range(12, ws_d.max_row + 1):
                ms = str(ws_d.cell(row=r, column=2).value or '').strip().upper()
                if ms != section.upper():
                    continue
                masters.append({
                    'row': r,
                    'desc': str(ws_d.cell(row=r, column=1).value or ''),
                    'po': str(ws_d.cell(row=r, column=7).value or ''),
                })
            used = set()
            for row in rows:
                best = {'score': 0, 'mr': None}
                for mr in masters:
                    if mr['row'] in used:
                        continue
                    s = match_score(row, mr)
                    if s > best['score']:
                        best = {'score': s, 'mr': mr}
                if best['score'] >= 4 and best['mr']:
                    ws.cell(row=best['mr']['row'], column=MASTER_152_OUTPUT_COL + 1).value = row['qty']
                    used.add(best['mr']['row'])
                    matched152 += 1

    # Save to bytes
    out_bytes = io.BytesIO()
    wb.save(out_bytes)
    out_bytes.seek(0)
    wb.close()
    wb_data.close()

    return out_bytes.getvalue(), {'matched153': matched153, 'matched152': matched152, 'unmatched': unmatched, 'total': sum(len(v) for v in sample_data.values())}


# ============================================================
# HTML TEMPLATE (embedded)
# ============================================================
def get_html():
    with open('index_flask.html', 'r', encoding='utf-8') as f:
        return f.read()


# ============================================================
# ROUTES
# ============================================================
@app.route('/')
def index():
    return get_html()


@app.route('/api/consolidate', methods=['POST'])
def api_consolidate():
    try:
        # Parse inputs
        master_file = request.files.get('master')
        sample_files = request.files.getlist('samples')
        month = request.form.get('month', 'May')
        password = request.form.get('password', '')

        if not master_file:
            return jsonify({'error': 'Thiếu file master'}), 400
        if not sample_files or len(sample_files) == 0:
            return jsonify({'error': 'Thiếu file sample'}), 400

        # Decrypt master
        master_bytes = master_file.read()
        try:
            master_bytes = decrypt_excel(master_bytes, password if password else 'sp')
        except ValueError as e:
            return jsonify({'error': str(e)}), 400

        # Read all samples
        all_data = {}
        sample_report = []
        for sf in sample_files:
            fbytes = sf.read()
            data = read_sample_data(fbytes)
            sec_count = len(data)
            row_count = sum(len(v) for v in data.values())
            sample_report.append({'name': sf.filename, 'sections': sec_count, 'rows': row_count})
            for s, rows in data.items():
                all_data.setdefault(s, []).extend(rows)

        if not all_data:
            return jsonify({'error': 'Không có dữ liệu output nào trong file sample!'}), 400

        # Consolidate
        result_bytes, stats = consolidate(master_bytes, all_data, month)

        # Save result temporarily
        out_name = f"Detail_Spare_parts_{month}26_CONSOLIDATED.xlsx"
        out_path = os.path.join(app.config['UPLOAD_FOLDER'], out_name)
        with open(out_path, 'wb') as f:
            f.write(result_bytes)

        return jsonify({
            'success': True,
            'filename': out_name,
            'stats': stats,
            'sample_report': sample_report,
            'download_url': f'/api/download/{out_name}'
        })

    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500


@app.route('/api/download/<filename>')
def api_download(filename):
    path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    if not os.path.exists(path):
        return jsonify({'error': 'File không tồn tại'}), 404
    return send_file(path, as_attachment=True, download_name=filename)


# ============================================================
# MAIN
# ============================================================
if __name__ == '__main__':
    port = int(os.environ.get('PORT', 8787))
    app.run(host='0.0.0.0', port=port)
